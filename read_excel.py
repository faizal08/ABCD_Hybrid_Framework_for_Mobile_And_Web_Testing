import openpyxl

# Load the workbook
wb = openpyxl.load_workbook('testingwe12.xlsx')
ws = wb['Sheet4']

# Get headers
headers = [cell.value for cell in ws[1]]
print("Column Headers:")
for i, header in enumerate(headers):
    print(f"  Column {i}: {header}")

print("\n" + "="*80)
print("Searching for Requirement ID: SA_TS_3")
print("="*80 + "\n")

# Find rows with SA_TS_3
found_rows = []
for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    if any(cell and 'SA_TS_3' in str(cell) for cell in row):
        found_rows.append((i, row))

if found_rows:
    print(f"Found {len(found_rows)} row(s) with SA_TS_3:\n")
    for row_num, row_data in found_rows:
        print(f"Row {row_num}:")
        for i, (header, value) in enumerate(zip(headers, row_data)):
            if value is not None and str(value).strip():
                print(f"  {header}: {value}")
        print()
else:
    print("No rows found with SA_TS_3")
    print("\nShowing first 5 rows to help identify the data:")
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=6, values_only=True), start=2):
        print(f"\nRow {i}:")
        for header, value in zip(headers, row):
            if value is not None:
                print(f"  {header}: {value}")
